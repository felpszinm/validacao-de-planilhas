from openpyxl import load_workbook
from datetime import datetime
import os

# Caminho do arquivo aonde está a pasta.xlsm.
FILE_PATH = os.path.join("Planilha", "Planilha.xlsx")

# Abre a planilha.
arquivo = load_workbook(FILE_PATH, data_only=True) # data_only=True (pega os valores já calculados das formulas do excel.)

# Abro o arquivo com as formulas, para salvar sem perde-las.
arquivo_com_formulas = load_workbook(FILE_PATH)

# Função para conversão de data americana para brasileira.
def conversao_para_data_brasileira(data_padrao):
    if data_padrao is None:
        return None
    try:
        # Tenta transformar a data americana em data brasileira:
        data_brasileira = data_padrao.strftime('%d/%m/%Y')
        return data_brasileira
    except ValueError:
        # Se caso der erro, retorna 'Data Inválida'
        return 'Data inválida!'


# Variáveis:
resumo = arquivo['RESUMO']
controle_rentals = arquivo['CONTROLE_RENTALS']
controle_rentals_com_formulas = arquivo_com_formulas['CONTROLE_RENTALS']

geral = arquivo['DB_GERAL']
rotas = arquivo['DB_ROTAS']
manutencao = arquivo['DB_MANU']
placas = arquivo['DB_PLACAS']
data_referencia = geral['A3'].value

# Verifica se a data de referencia é válida.
if not isinstance(data_referencia, datetime):
    raise ValueError('A data de referência é inválida!')

placas_e_datas_list = []

# Para cada linha dentro de DB_GERAL
for linha in range(2, rotas.max_row):

    #Se caso a linha de rota for vazia, ele para de adicionar placas e datas   
    if rotas[f'C{linha}'].value is None:
        print('Sua lista de placas chegou ao fim!')
        break
    
    # Pega os valores das abas C/D do excel, e já faz a conversão para data brasileira.
    data_padrao = rotas[f'C{linha}'].value
    data_brasileira = conversao_para_data_brasileira(data_padrao)
    placa = rotas[f'D{linha}'].value

    # Se a data já formatada não estiver vazia, ele adiciona a lista.
    if data_brasileira is not None:
        placas_e_datas_list.append([data_brasileira, placa])

# Para cada placa|data que está dentro da lista de placas e datas:        
idx_linhas = 3      
for placa_na_lista in placas_e_datas_list:

    # Pega os valores de cada lista de placas separadamentes:
    idx_datas = 0
    idx_placas = 1   
    data = placa_na_lista[idx_datas]
    placa = placa_na_lista[idx_placas]

    # Adiciona esses valores na planilha:
    controle_rentals_com_formulas.cell(idx_linhas, column=2).value = data
    controle_rentals_com_formulas.cell(idx_linhas, column=3).value = placa
    
    # Vai verificar se essa placa está em manutenção: (Valor padrão -> NÃO)
    verificar_manutencao = 'NÃO'
    for manu_linhas in range(2, manutencao.max_row + 1):
        placa_manu = manutencao.cell(manu_linhas, column=1).value
        data_entrada = manutencao.cell(manu_linhas, column=4).value
        data_saida = manutencao.cell(manu_linhas, column=5).value

        # Se a data_saida for string, ele transforma em datetime:
        if isinstance(data_saida, str):
            try:
                data_saida = datetime.strptime(data_saida, '%d/%m/%Y')
            except:
                data_saida = None
                
        # Se a placa for igual a placa dentro da manutenção, ele vai alterar 'verificar_manutencao' para 'SIM':
        if placa_manu == placa:
            if isinstance(data_entrada, datetime):
                if data_referencia >= data_entrada and (data_saida is None or data_referencia <= data_saida):
                    verificar_manutencao = 'SIM'
                    break
    
    # Adiciona o valor na planilha:
    controle_rentals_com_formulas.cell(idx_linhas, column=9).value = verificar_manutencao
    idx_linhas += 1

# Por fim salva a planilha.
arquivo_com_formulas.save(FILE_PATH)
print('Seu programa foi concluido com sucesso!✅')

# //TODO: Criar um verificador para se caso a placa tiver '-', retirar o traço e salva-la sem.